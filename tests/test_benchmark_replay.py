"""Benchmark replay logic tests (PRD §9).

The benchmark replay is the #1 product differentiator (94.6% path-match on
held-out 2025). These tests prove the replay logic is correct, deterministic,
and produces expected accuracy metrics on synthetic data with known outcomes.
"""
from __future__ import annotations

import math
import pandas as pd
import pytest

from benchmark.replay import replay
from benchmark.score import score
from benchmark.normalize import load

# ── helpers ──────────────────────────────────────────────────────────────────

MIN_HEADROOM = 50
CAP = 0.20


def _df(rows: list[dict]) -> pd.DataFrame:
    """Build a synthetic DataFrame matching the normalize.py output schema. All
    columns that replay.py and score.py touch are included; extra columns are
    silently ignored by the functions under test."""
    return pd.DataFrame(rows)


def _row(*, salary: float, emi: float, arrears: float, approved_path: str,
          add_prem: float = 0, add_months: float = 0, year: int = 2025) -> dict:
    """Single row factory. Most tests override the relevant fields after
    construction; this keeps the boilerplate in one place.

    The returned dict includes 'add_prem' because score.py references that
    column — it is normally added by normalize.load() before replay() is called.
    """
    return {
        "CURRENT_SALARY": salary,
        "CURRENT_EMI_AMT": emi,
        "OVER_DUE_AMT": arrears,
        "APPROVED_REQUEST_TYPE": approved_path,
        "ADDITIONAL_PREMIUM": add_prem,
        "ADDITIONAL_MONTHS": add_months,
        "year": year,
        "add_prem": add_prem,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# S1 — path-match accuracy: perfect, partial, zero
# ═══════════════════════════════════════════════════════════════════════════════

class TestPathMatchAccuracy:
    """Scenario S1: replay() correctly classifies every row and score() reports
    the right path-match accuracy for synthetic data."""

    def test_perfect_accuracy_100_pct(self):
        """All headroom > 50 → UPDATE_INSTALLMENT, and all approved as UPDATE."""
        df = _df([
            _row(salary=10000, emi=1000, arrears=5000,  # headroom = 1000
                 approved_path="UPDATE_INSTALLMENT"),
            _row(salary=8000, emi=500, arrears=2000,     # headroom = 1100
                 approved_path="UPDATE_INSTALLMENT"),
            _row(salary=5000, emi=200, arrears=1000,     # headroom = 800
                 approved_path="UPDATE_INSTALLMENT"),
            _row(salary=3000, emi=2500, arrears=1000,    # headroom = -1900 → TRANSFER
                 approved_path="TRANSFER_ARREARS"),
            _row(salary=2000, emi=1800, arrears=500,     # headroom = -1400 → TRANSFER
                 approved_path="TRANSFER_ARREARS"),
        ])
        result = score(replay(df))
        assert result["path_match_accuracy"] == 1.0
        assert result["n"] == 5

    def test_partial_accuracy_60_pct(self):
        """3 of 5 match → 0.6."""
        df = _df([
            # Match (headroom > 50, labeled UPDATE)
            _row(salary=10000, emi=1000, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT"),    # ✓ match
            _row(salary=8000, emi=500, arrears=2000,
                 approved_path="UPDATE_INSTALLMENT"),    # ✓ match
            _row(salary=5000, emi=200, arrears=1000,
                 approved_path="UPDATE_INSTALLMENT"),    # ✓ match
            # Mismatch (headroom > 50 but labeled TRANSFER)
            _row(salary=6000, emi=500, arrears=3000,
                 approved_path="TRANSFER_ARREARS"),      # ✗ pred=UPDATE
            # Mismatch (headroom <= 50 but labeled UPDATE)
            _row(salary=2500, emi=2500, arrears=500,
                 approved_path="UPDATE_INSTALLMENT"),    # ✗ pred=TRANSFER
        ])
        result = score(replay(df))
        assert result["path_match_accuracy"] == pytest.approx(0.6, abs=0.01)
        assert result["n"] == 5

    def test_zero_accuracy_headroom_mismatch(self):
        """All headroom > 50 but all labeled TRANSFER → 0.0."""
        df = _df([
            _row(salary=10000, emi=1000, arrears=5000,
                 approved_path="TRANSFER_ARREARS"),      # pred=UPDATE, label=TRANSFER
            _row(salary=8000, emi=500, arrears=2000,
                 approved_path="TRANSFER_ARREARS"),
            _row(salary=5000, emi=200, arrears=1000,
                 approved_path="TRANSFER_ARREARS"),
        ])
        result = score(replay(df))
        assert result["path_match_accuracy"] == 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# S2 — headroom boundary: exactly MIN_HEADROOM (50)
# ═══════════════════════════════════════════════════════════════════════════════

class TestHeadroomBoundary:
    """Scenario S2: the boundary condition at exactly MIN_HEADROOM (50)."""

    def test_headroom_equals_50_is_transfer(self):
        """Headroom == 50 is NOT > 50, so it must classify as TRANSFER_ARREARS."""
        df = _df([
            # headroom = 0.20 * 10000 - 1950 = 2000 - 1950 = 50 → TRANSFER
            _row(salary=10000, emi=1950, arrears=5000,
                 approved_path="TRANSFER_ARREARS"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_path"] == "TRANSFER_ARREARS"

    def test_headroom_51_is_update(self):
        """Headroom == 51 is > 50, so UPDATE_INSTALLMENT."""
        df = _df([
            # headroom = 0.20 * 10000 - 1949 = 2000 - 1949 = 51 → UPDATE
            _row(salary=10000, emi=1949, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_path"] == "UPDATE_INSTALLMENT"

    def test_headroom_negative_is_transfer(self):
        """Negative headroom (EMI > cap) is always TRANSFER."""
        df = _df([
            _row(salary=5000, emi=5000, arrears=1000,
                 approved_path="TRANSFER_ARREARS"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_path"] == "TRANSFER_ARREARS"
        # pred_prem must be 0 (floor of negative headroom = 0)
        assert result.iloc[0]["pred_prem"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# S3 — premium and months computation
# ═══════════════════════════════════════════════════════════════════════════════

class TestPremiumAndMonths:
    """Scenario S3: the financial computation inside replay() is correct."""

    def test_update_premium_is_floor_of_headroom(self):
        """pred_prem = floor(headroom). Example: headroom 1234 → 1234."""
        df = _df([
            _row(salary=10000, emi=500, arrears=5000,  # headroom=1500
                 approved_path="UPDATE_INSTALLMENT"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_prem"] == 1500

    def test_update_months_ceil_of_arrears_over_premium(self):
        """pred_months = ceil(arrears / pred_prem). Arrears=5000, prem=1500 → 4."""
        df = _df([
            _row(salary=10000, emi=500, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_months"] == math.ceil(5000 / 1500)  # 4

    def test_zero_arrears_means_zero_months(self):
        """Zero arrears → zero additional months. No division error."""
        df = _df([
            _row(salary=10000, emi=500, arrears=0,
                 approved_path="UPDATE_INSTALLMENT"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_months"] == 0

    def test_transfer_path_zero_premium(self):
        """TRANSFER path: pred_prem = 0, pred_months = 0."""
        df = _df([
            _row(salary=3000, emi=3000, arrears=5000,
                 approved_path="TRANSFER_ARREARS"),
        ])
        result = replay(df)
        assert result.iloc[0]["pred_prem"] == 0
        assert result.iloc[0]["pred_months"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# S4 — 20% compliance
# ═══════════════════════════════════════════════════════════════════════════════

class TestTwentyPctCompliance:
    """Scenario S4: every UPDATE plan must be within the 20% cap (threshold
    0.205 in score.py accounts for floating-point rounding)."""

    def test_plan_below_cap_is_compliant(self):
        """Deduction rate well under 20% → compliant (1.0)."""
        df = _df([
            _row(salary=10000, emi=500, arrears=1000,
                 approved_path="UPDATE_INSTALLMENT"),
            # new_total = 500 + 1500 = 2000, ded_ratio = 0.20
            _row(salary=5000, emi=200, arrears=500,
                 approved_path="UPDATE_INSTALLMENT"),
            # new_total = 200 + 800 = 1000, ded_ratio = 0.20
        ])
        result = score(replay(df))
        # floor(1500)=1500, fd=2000/10000=0.20
        # floor(800)=800, fd=1000/5000=0.20
        assert result["twenty_pct_compliance_update_plans"] == 1.0

    def test_plan_at_exact_cap_passes(self):
        """Deduction exactly at 0.20 → must pass (1.0)."""
        df = _df([
            _row(salary=10000, emi=2000, arrears=1000,  # headroom=0 → TRANSFER
                 approved_path="UPDATE_INSTALLMENT"),    # intentionally mismatched
        ])
        # headroom=0, so pred_path=TRANSFER → not in pu (UPDATE predictions)
        # So this test needs actual UPDATE predictions
        df2 = _df([
            _row(salary=10000, emi=1800, arrears=2000,  # headroom=200
                 approved_path="UPDATE_INSTALLMENT"),
        ])
        r = replay(df2)
        # new_total = 1800 + 200 = 2000, ded_ratio = 2000/10000 = 0.20
        assert r.iloc[0]["pred_ded_ratio"] == pytest.approx(0.20, abs=0.001)
        result = score(r)
        assert result["twenty_pct_compliance_update_plans"] == 1.0


# ═══════════════════════════════════════════════════════════════════════════════
# S5 — multi-year scoring and held-out validation
# ═══════════════════════════════════════════════════════════════════════════════

class TestMultiYearScoring:
    """Scenario S5: score() correctly filters by year for hold-out validation."""

    def test_all_years_accuracy(self):
        """score(df) without year filter uses all rows."""
        df = _df([
            _row(salary=10000, emi=1000, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT", year=2023),
            _row(salary=8000, emi=500, arrears=2000,
                 approved_path="UPDATE_INSTALLMENT", year=2024),
            _row(salary=3000, emi=2500, arrears=1000,
                 approved_path="TRANSFER_ARREARS", year=2025),
        ])
        result = score(replay(df))
        assert result["n"] == 3
        assert result["path_match_accuracy"] == 1.0

    def test_held_out_year_filter(self):
        """score(df, 2025) returns only 2025 rows."""
        df = _df([
            _row(salary=10000, emi=1000, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT", year=2023),
            _row(salary=8000, emi=500, arrears=2000,
                 approved_path="UPDATE_INSTALLMENT", year=2024),
            _row(salary=3000, emi=2500, arrears=1000,
                 approved_path="TRANSFER_ARREARS", year=2025),
        ])
        result = score(replay(df), year=2025)
        assert result["n"] == 1
        assert result["path_match_accuracy"] == 1.0

    def test_held_out_different_accuracy(self):
        """Different years can produce different accuracy (proves filtering works)."""
        df = _df([
            # 2023: all match
            _row(salary=10000, emi=1000, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT", year=2023),
            # 2025: mismatch
            _row(salary=3000, emi=2500, arrears=1000,
                 approved_path="UPDATE_INSTALLMENT", year=2025),
        ])
        all_years = score(replay(df))
        held_out = score(replay(df), year=2025)
        assert all_years["path_match_accuracy"] == 0.5
        assert held_out["path_match_accuracy"] == 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# S6 — deterministic across runs
# ═══════════════════════════════════════════════════════════════════════════════

class TestDeterminism:
    """Scenario S6: replay() is deterministic — same input, same output."""

    def test_replay_is_deterministic(self):
        """Two consecutive calls with the same DataFrame produce identical pred_path."""
        df = _df([
            _row(salary=10000, emi=1000, arrears=5000,
                 approved_path="UPDATE_INSTALLMENT"),
            _row(salary=3000, emi=2500, arrears=1000,
                 approved_path="TRANSFER_ARREARS"),
        ])
        r1 = replay(df)
        r2 = replay(df)
        assert (r1["pred_path"] == r2["pred_path"]).all()


# ═══════════════════════════════════════════════════════════════════════════════
# S7 — normalize.py loads data correctly (integration smoke test)
# ═══════════════════════════════════════════════════════════════════════════════

class TestNormalize:
    """Scenario S7: normalize.load() handles the real workbook OR a synthetic
    workbook with the expected schema. This test creates a small synthetic xlsx
    to prove the pipeline works end-to-end."""

    @pytest.fixture
    def synthetic_xlsx(self, tmp_path):
        """Create a minimal synthetic workbook matching the 3-sheet schema."""
        import openpyxl
        path = tmp_path / "synthetic_benchmark.xlsx"
        wb = openpyxl.Workbook()
        # Remove default sheet, create year sheets
        for year, rows in [
            ("2023", [
                {"CURRENT_SALARY": 10000, "OVER_DUE_AMT": 5000,
                 "CURRENT_EMI_AMT": 1000, "NEW_EMI_AMT": 2000,
                 "ADDITIONAL_MONTHS": 4, "ADDITIONAL_PREMIUM": 1000,
                 "APPROVED_REQUEST_TYPE": "UPDATE_INSTALLMENT"},
            ]),
            ("2024", [
                {"CURRENT_SALARY": 8000, "OVER_DUE_AMT": 2000,
                 "CURRENT_EMI_AMT": 500, "NEW_EMI_AMT": 1300,
                 "ADDITIONAL_MONTHS": 2, "ADDITIONAL_PREMIUM": 800,
                 "APPROVED_REQUEST_TYPE": "UPDATE_INSTALLMENT"},
            ]),
            ("2025", [
                {"CURRENT_SALARY": 3000, "OVER_DUE_AMT": 1000,
                 "CURRENT_EMI_AMT": 2500, "NEW_EMI_AMT": 2500,
                 "ADDITIONAL_MONTHS": 0, "ADDITIONAL_PREMIUM": 0,
                 "APPROVED_REQUEST_TYPE": "TRANSFER_ARREARS"},
                {"CURRENT_SALARY": 5000, "OVER_DUE_AMT": 10000,
                 "CURRENT_EMI_AMT": 4800, "NEW_EMI_AMT": 4800,
                 "ADDITIONAL_MONTHS": 0, "ADDITIONAL_PREMIUM": 0,
                 "APPROVED_REQUEST_TYPE": "TRANSFER_ARREARS"},
            ]),
        ]:
            ws = wb.create_sheet(title=year)
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for row in rows:
                    ws.append([row[h] for h in headers])
        # Remove the auto-created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(path)
        return path

    def test_load_synthetic_workbook(self, synthetic_xlsx):
        """Smoke test: normalize.load() can read a properly structured workbook."""
        df = load(str(synthetic_xlsx))
        assert len(df) == 4  # 1 + 1 + 2 rows across 3 sheets
        assert list(df.columns) == [
            "CURRENT_SALARY", "OVER_DUE_AMT", "CURRENT_EMI_AMT", "NEW_EMI_AMT",
            "ADDITIONAL_MONTHS", "ADDITIONAL_PREMIUM",
            "APPROVED_REQUEST_TYPE", "year", "add_prem",
        ]
        assert set(df["year"]) == {2023, 2024, 2025}

    def test_replay_scores_synthetic_workbook(self, synthetic_xlsx):
        """End-to-end: load → replay → score produces plausible metrics."""
        df = load(str(synthetic_xlsx))
        result = score(replay(df))
        # All 4 rows should have a pred_path
        assert result["n"] == 4
        # The 2 TRANSFER rows should match (both headroom <= 50)
        # The 2 UPDATE rows should match (both headroom > 50)
        assert result["path_match_accuracy"] == 1.0
